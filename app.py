import streamlit as st
import pandas as pd
import openpyxl
import re
import requests
import os
from io import BytesIO

# ───────────────────────── FILE PATHS ─────────────────────────
DEFAULT_TEMPLATE = "sku-template (4).xlsx"
FALLBACK_UPLOADED_TEMPLATE = "/mnt/data/output_template (62).xlsx"

if os.path.exists(FALLBACK_UPLOADED_TEMPLATE):
    TEMPLATE_PATH = FALLBACK_UPLOADED_TEMPLATE
else:
    TEMPLATE_PATH = DEFAULT_TEMPLATE

# ═══════════════════════════════════════════════════════════════
#  BATCH ID  —  Stored in Google Sheets, shared across all users
#
#  HOW IT WORKS:
#    A Google Apps Script Web App acts as a tiny API sitting in
#    front of your sheet. It handles GET (read) and POST (write).
#    Because it's deployed as "Anyone, even anonymous" it needs
#    zero credentials — just an HTTP call.
#
#  ONE-TIME SETUP (~3 minutes):
#  ─────────────────────────────
#  1. Open your Google Sheet:
#     https://docs.google.com/spreadsheets/d/1oxtgaZmfJseMoiOlqGRkm2pWSQoga5Ys-jcDZGTUFEM
#  2. Put the number  1  in cell A1  (this is your starting BatchID)
#  3. Click  Extensions → Apps Script
#  4. Delete any existing code and paste this ENTIRE script:
#
# ┌──────────────────────────────────────────────────────────────
# │ var SHEET_ID = "1oxtgaZmfJseMoiOlqGRkm2pWSQoga5Ys-jcDZGTUFEM";
# │
# │ function doGet(e) {
# │   var sheet = SpreadsheetApp.openById(SHEET_ID).getSheets()[0];
# │   var val = sheet.getRange("A1").getValue();
# │   return ContentService
# │     .createTextOutput(JSON.stringify({ batch_id: val }))
# │     .setMimeType(ContentService.MimeType.JSON);
# │ }
# │
# │ function doPost(e) {
# │   var data  = JSON.parse(e.postData.contents);
# │   var sheet = SpreadsheetApp.openById(SHEET_ID).getSheets()[0];
# │   sheet.getRange("A1").setValue(data.next_id);
# │   return ContentService
# │     .createTextOutput(JSON.stringify({ ok: true, saved: data.next_id }))
# │     .setMimeType(ContentService.MimeType.JSON);
# │ }
# └──────────────────────────────────────────────────────────────
#
#  5. Click  Deploy → New deployment
#     • Type: Web app
#     • Execute as: Me
#     • Who has access: Anyone
#  6. Click Deploy → copy the Web App URL
#  7. Paste that URL as the value of  APPS_SCRIPT_URL  below
# ═══════════════════════════════════════════════════════════════

APPS_SCRIPT_URL = "https://script.google.com/macros/s/AKfycbxiCe1IVsghaaFa4zJvA-YuCowvvT3JzLZag1IAp9B8MFGk6w8hI4aBpoB_WsqWkbbLPg/exec"

# ── Optionally load from Streamlit secrets (for cloud deployments) ──
try:
    if not APPS_SCRIPT_URL:
        APPS_SCRIPT_URL = st.secrets["APPS_SCRIPT_URL"]
except Exception:
    pass

# ── Local file fallback (single machine, used if URL not set) ──
_FALLBACK_FILE = "batch_id_counter.json"

def _local_read() -> int:
    import json
    if os.path.exists(_FALLBACK_FILE):
        try:
            with open(_FALLBACK_FILE) as f:
                return int(json.load(f).get("v", 1))
        except Exception:
            pass
    return 1

def _local_write(v: int):
    import json
    tmp = _FALLBACK_FILE + ".tmp"
    with open(tmp, "w") as f:
        json.dump({"v": v}, f)
    os.replace(tmp, _FALLBACK_FILE)

# ── Google Sheets read / write via Apps Script ─────────────────
def _remote_read() -> int:
    r = requests.get(APPS_SCRIPT_URL, timeout=10)
    r.raise_for_status()
    return int(r.json()["batch_id"])

def _remote_write(next_id: int):
    r = requests.post(APPS_SCRIPT_URL, json={"next_id": next_id}, timeout=10)
    r.raise_for_status()

# ── Public helpers ─────────────────────────────────────────────
def peek_next_batch_id() -> int:
    """Read current BatchID without consuming it (for UI display)."""
    if APPS_SCRIPT_URL:
        try:
            return _remote_read()
        except Exception:
            pass
    return _local_read()

def get_and_increment_batch_id() -> int:
    """
    Atomically claim the current BatchID and advance the counter.
    Returns the claimed BatchID.
    """
    if APPS_SCRIPT_URL:
        try:
            current = _remote_read()
            _remote_write(current + 1)
            return current
        except Exception as e:
            st.warning(f"⚠️ Google Sheets BatchID unavailable ({e}). Using local counter.")
    # Fallback
    current = _local_read()
    _local_write(current + 1)
    return current


# ╭───────────────── NORMALISERS & HELPERS ─────────────────╮
def norm(s) -> str:
    if pd.isna(s):
        return ""
    return "".join(str(s).split()).lower()

def clean_header(header) -> str:
    if pd.isna(header):
        return ""
    header_str = str(header)
    header_str = re.sub(r"[^0-9A-Za-z ]+", " ", header_str)
    header_str = re.sub(r"\s+", " ", header_str).strip()
    return header_str

IMAGE_EXT_RE = re.compile(r"(?i)\.(jpe?g|png|gif|bmp|webp|tiff?)$")
IMAGE_KEYWORDS = {"image", "img", "picture", "photo", "thumbnail", "thumb", "hero", "front", "back", "url"}

def is_image_column(col_header_norm: str, series: pd.Series) -> bool:
    header_hit = any(k in col_header_norm for k in IMAGE_KEYWORDS)
    sample = series.dropna().astype(str).head(20)
    ratio = sample.str.contains(IMAGE_EXT_RE).mean() if not sample.empty else 0.0
    return header_hit or ratio >= 0.30

def dedupe_columns(columns):
    seen = {}
    result = []
    for col in columns:
        col_str = str(col) if not pd.isna(col) else "Unnamed"
        if col_str in seen:
            seen[col_str] += 1
            result.append(f"{col_str}_{seen[col_str]}")
        else:
            seen[col_str] = 0
            result.append(col_str)
    return result
# ╰───────────────────────────────────────────────────────────╯

MARKETPLACE_ID_MAP = {
    "Amazon":   ("Parent SKU", "SKU"),
    "Myntra":   ("styleId", "styleGroupId"),
    "Ajio":     ("*Item SKU", "*Style Code"),
    "Flipkart": ("Seller SKU ID", "Style Code"),
    "TataCliq": ("Seller Article SKU", "*Style Code"),
    "Zivame":   ("Style Code", "SKU Code"),
    "Celio":    ("Style Code", "SKU Code"),
}

def find_column_by_name_like(src_df: pd.DataFrame, name: str):
    if not name:
        return None
    name = str(name).strip()
    for c in src_df.columns:
        if str(c).strip() == name:
            return c
    nname = norm(name)
    for c in src_df.columns:
        if norm(c) == nname:
            return c
    for c in src_df.columns:
        if nname in norm(c):
            return c
    return None

def read_input_to_df(input_file, marketplace, header_row=1, data_row=2, sheet_name=None):
    marketplace_configs = {
        "Amazon":   {"sheet": "Template", "header_row": 4, "data_row": 7,  "sheet_index": None},
        "Flipkart": {"sheet": None,       "header_row": 1, "data_row": 5,  "sheet_index": 2},
        "Myntra":   {"sheet": None,       "header_row": 3, "data_row": 4,  "sheet_index": 1},
        "Ajio":     {"sheet": None,       "header_row": 2, "data_row": 3,  "sheet_index": 2},
        "TataCliq": {"sheet": None,       "header_row": 4, "data_row": 6,  "sheet_index": 0},
        "General":  {"sheet": None,       "header_row": header_row, "data_row": data_row, "sheet_index": 0}
    }
    config = marketplace_configs.get(marketplace, marketplace_configs["General"])

    if marketplace == "General" and sheet_name:
        xl = pd.ExcelFile(input_file)
        temp_df = xl.parse(sheet_name, header=None)
        header_idx = header_row - 1
        data_idx = data_row - 1
        headers = temp_df.iloc[header_idx].tolist()
        src_df = temp_df.iloc[data_idx:].copy()
        src_df.columns = dedupe_columns(headers)
        src_df.reset_index(drop=True, inplace=True)

    elif config["sheet"] is not None:
        xl = pd.ExcelFile(input_file)
        temp_df = xl.parse(config["sheet"], header=None)
        header_idx = config["header_row"] - 1
        data_idx = config["data_row"] - 1
        headers = temp_df.iloc[header_idx].tolist()
        src_df = temp_df.iloc[data_idx:].copy()
        src_df.columns = dedupe_columns(headers)
        src_df.reset_index(drop=True, inplace=True)
        if marketplace == "Amazon":
            parentage_col = find_column_by_name_like(src_df, "Parentage Level")
            if parentage_col:
                before = len(src_df)
                src_df = src_df[
                    src_df[parentage_col].astype(str).str.strip().str.lower() != "parent"
                ].copy()
                src_df.reset_index(drop=True, inplace=True)
                after = len(src_df)
                src_df.attrs["filtered_parent_rows"] = before - after
    else:
        xl = pd.ExcelFile(input_file)
        temp_df = xl.parse(xl.sheet_names[config["sheet_index"]], header=None)
        header_idx = config["header_row"] - 1
        data_idx = config["data_row"] - 1
        headers = temp_df.iloc[header_idx].tolist()
        src_df = temp_df.iloc[data_idx:].copy()
        src_df.columns = dedupe_columns(headers)
        src_df.reset_index(drop=True, inplace=True)

    src_df.dropna(axis=1, how='all', inplace=True)
    return src_df


def process_file(
    input_file,
    marketplace: str,
    selected_variant_col: str | None = None,
    selected_product_col: str | None = None,
    general_header_row: int = 1,
    general_data_row: int = 2,
    general_sheet_name: str | None = None,
):
    src_df = read_input_to_df(
        input_file, marketplace,
        header_row=general_header_row,
        data_row=general_data_row,
        sheet_name=general_sheet_name
    )

    # ── Claim BatchID FIRST (before any processing) ──────────────
    batch_id     = get_and_increment_batch_id()
    batch_id_str = str(batch_id)
    num_rows     = len(src_df)

    # auto-map every column
    columns_meta = []
    for col in src_df.columns:
        dtype = "imageurlarray" if is_image_column(norm(col), src_df[col]) else "string"
        columns_meta.append({"src": col, "out": col, "row3": "mandatory", "row4": dtype})

    # identify color/size
    color_cols = [col for col in src_df.columns if "color" in norm(col) or "colour" in norm(col)]
    size_cols  = [col for col in src_df.columns if "size" in norm(col)]

    option1_data = pd.Series([""] * num_rows, dtype=str)
    option2_data = pd.Series([""] * num_rows, dtype=str)
    if size_cols:
        option1_data = src_df[size_cols[0]].fillna('').astype(str).str.strip()
        if color_cols and color_cols[0] != size_cols[0]:
            option2_data = src_df[color_cols[0]].fillna('').astype(str).str.strip()
    elif color_cols:
        option2_data = src_df[color_cols[0]].fillna('').astype(str).str.strip()

    unique_opt1 = option1_data.replace("", pd.NA).dropna().unique().tolist()
    unique_opt2 = option2_data.replace("", pd.NA).dropna().unique().tolist()

    # load workbook
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws_vals  = wb["Values"]
    ws_types = wb["Types"]

    def first_empty_col(ws, header_rows=(1,)):
        for col_idx in range(1, 201):
            empty = True
            for r in header_rows:
                if ws.cell(row=r, column=col_idx).value not in (None, ""):
                    empty = False
                    break
            if empty:
                return col_idx
        return ws.max_column + 1

    vals_start_col  = first_empty_col(ws_vals,  header_rows=(1,))
    types_start_col = first_empty_col(ws_types, header_rows=(1, 2, 3, 4))

    # Write columns_meta
    for idx, meta in enumerate(columns_meta):
        vcol = vals_start_col  + idx
        tcol = types_start_col + idx
        header_display = clean_header(meta["out"])
        ws_vals.cell(row=1, column=vcol, value=header_display)
        for r_idx, value in enumerate(src_df[meta["src"]].tolist(), start=2):
            cell = ws_vals.cell(row=r_idx, column=vcol)
            if pd.isna(value):
                cell.value = None
            else:
                if str(meta["row4"]).lower() in ("string", "imageurlarray"):
                    cell.value = str(value)
                    cell.number_format = "@"
                else:
                    cell.value = value
        ws_types.cell(row=1, column=tcol, value=header_display)
        ws_types.cell(row=2, column=tcol, value=header_display)
        ws_types.cell(row=3, column=tcol, value=meta["row3"])
        ws_types.cell(row=4, column=tcol, value=meta["row4"])

    # Append Option1 & Option2
    opt1_vcol = vals_start_col  + len(columns_meta)
    opt2_vcol = opt1_vcol + 1
    ws_vals.cell(row=1, column=opt1_vcol, value="Option 1")
    ws_vals.cell(row=1, column=opt2_vcol, value="Option 2")
    for i, v in enumerate(option1_data.tolist(), start=2):
        ws_vals.cell(row=i, column=opt1_vcol, value=v if v else None)
    for i, v in enumerate(option2_data.tolist(), start=2):
        ws_vals.cell(row=i, column=opt2_vcol, value=v if v else None)

    opt1_tcol = types_start_col + len(columns_meta)
    opt2_tcol = opt1_tcol + 1
    ws_types.cell(row=1, column=opt1_tcol, value="Option 1")
    ws_types.cell(row=2, column=opt1_tcol, value="Option 1")
    ws_types.cell(row=3, column=opt1_tcol, value="non mandatory")
    ws_types.cell(row=4, column=opt1_tcol, value="select")
    ws_types.cell(row=1, column=opt2_tcol, value="Option 2")
    ws_types.cell(row=2, column=opt2_tcol, value="Option 2")
    ws_types.cell(row=3, column=opt2_tcol, value="non mandatory")
    ws_types.cell(row=4, column=opt2_tcol, value="select")
    for i, val in enumerate(unique_opt1, start=5):
        ws_types.cell(row=i, column=opt1_tcol, value=val)
    for i, val in enumerate(unique_opt2, start=5):
        ws_types.cell(row=i, column=opt2_tcol, value=val)

    # Append variantId & productId
    def append_id_columns(variant_series, product_series):
        has_var  = variant_series is not None and variant_series.replace("", pd.NA).dropna().shape[0] > 0
        has_prod = product_series is not None and product_series.replace("", pd.NA).dropna().shape[0] > 0
        if not (has_var or has_prod):
            return
        after_written_vals  = vals_start_col  + len(columns_meta) + 2
        after_written_types = types_start_col + len(columns_meta) + 2
        cur_v = after_written_vals
        cur_t = after_written_types
        if has_var:
            ws_vals.cell(row=1, column=cur_v, value="variantId")
            for i, v in enumerate(variant_series.tolist(), start=2):
                cell = ws_vals.cell(row=i, column=cur_v, value=v if v else None)
                cell.number_format = "@"
            ws_types.cell(row=1, column=cur_t, value="variantId")
            ws_types.cell(row=2, column=cur_t, value="variantId")
            ws_types.cell(row=3, column=cur_t, value="mandatory")
            ws_types.cell(row=4, column=cur_t, value="string")
            cur_v += 1
            cur_t += 1
        if has_prod:
            ws_vals.cell(row=1, column=cur_v, value="productId")
            for i, v in enumerate(product_series.tolist(), start=2):
                cell = ws_vals.cell(row=i, column=cur_v, value=v if v else None)
                cell.number_format = "@"
            ws_types.cell(row=1, column=cur_t, value="productId")
            ws_types.cell(row=2, column=cur_t, value="productId")
            ws_types.cell(row=3, column=cur_t, value="mandatory")
            ws_types.cell(row=4, column=cur_t, value="string")

    if marketplace == "General":
        variant_series = None
        product_series = None
        if selected_variant_col and selected_variant_col != "(none)":
            if selected_variant_col in src_df.columns:
                variant_series = src_df[selected_variant_col].fillna("").astype(str)
        if selected_product_col and selected_product_col != "(none)":
            if selected_product_col in src_df.columns:
                product_series = src_df[selected_product_col].fillna("").astype(str)
        append_id_columns(variant_series, product_series)
    else:
        mapping = MARKETPLACE_ID_MAP.get(marketplace, None)
        if mapping:
            prod_src_name, var_src_name = mapping
            prod_col = find_column_by_name_like(src_df, prod_src_name)
            var_col  = find_column_by_name_like(src_df, var_src_name)
            prod_series = src_df[prod_col].fillna("").astype(str) if prod_col else None
            var_series  = src_df[var_col].fillna("").astype(str)  if var_col  else None
            append_id_columns(var_series, prod_series)

    # ── BatchID column — always the very last column ──────────────
    bv_col = first_empty_col(ws_vals, header_rows=(1,))
    ws_vals.cell(row=1, column=bv_col, value="BatchID")
    for r in range(2, num_rows + 2):
        cell = ws_vals.cell(row=r, column=bv_col, value=batch_id_str)
        cell.number_format = "@"

    bt_col = first_empty_col(ws_types, header_rows=(1, 2, 3, 4))
    ws_types.cell(row=1, column=bt_col, value="BatchID")
    ws_types.cell(row=2, column=bt_col, value="BatchID")
    ws_types.cell(row=3, column=bt_col, value="non mandatory")
    ws_types.cell(row=4, column=bt_col, value="string")
    # ─────────────────────────────────────────────────────────────

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, batch_id


# ───────────────────────── STREAMLIT UI ─────────────────────────
st.set_page_config(page_title="SKU Template Automation", layout="wide")
st.title("Rubick OS Template Conversion")

# Warn if Apps Script URL not configured
if not APPS_SCRIPT_URL:
    st.warning(
        "⚠️ **BatchID Google Sheets sync not configured.**  \n"
        "Set `APPS_SCRIPT_URL` in the code (or `secrets.toml`) after completing the "
        "one-time Apps Script setup described in the code comments.  \n"
        "Until then, BatchID falls back to a local file counter."
    )

# Show next BatchID (read-only peek)
next_id = peek_next_batch_id()
st.info(f"📦 Next BatchID to be assigned: **{next_id}**")

if os.path.exists(TEMPLATE_PATH):
    st.info(f"Using template: {os.path.basename(TEMPLATE_PATH)}")
    try:
        with open(TEMPLATE_PATH, "rb") as f:
            st.download_button(
                "Download current template (for reference)",
                data=f.read(),
                file_name=os.path.basename(TEMPLATE_PATH),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    except Exception:
        pass

marketplace_options = ["General", "Amazon", "Flipkart", "Myntra", "Ajio", "TataCliq", "Zivame", "Celio"]
marketplace_type = st.selectbox("Select Template Type", marketplace_options)

general_header_row = 1
general_data_row   = 2
if marketplace_type == "General":
    st.info("Callout: Leave blank to use defaults — Header row = 1, Data row = 2.")
    col_h, col_d = st.columns(2)
    with col_h:
        _hr = st.text_input("Header row (1-indexed)", value="", placeholder="Default: 1")
    with col_d:
        _dr = st.text_input("Data row (1-indexed)", value="", placeholder="Default: 2")
    try:
        general_header_row = int(_hr.strip()) if _hr.strip() else 1
    except ValueError:
        st.error("Header row must be a number.")
        general_header_row = 1
    try:
        general_data_row = int(_dr.strip()) if _dr.strip() else 2
    except ValueError:
        st.error("Data row must be a number.")
        general_data_row = 2

input_file = st.file_uploader("Upload Input Excel File", type=["xlsx", "xls", "xlsm"])

selected_variant_col = "(none)"
selected_product_col = "(none)"

if input_file:
    selected_sheet = None
    if marketplace_type == "General":
        try:
            xl     = pd.ExcelFile(input_file)
            sheets = xl.sheet_names
            selected_sheet = st.selectbox("Select sheet", sheets)
        except Exception as e:
            st.error(f"Failed to read sheets from uploaded file: {e}")
            selected_sheet = None

    try:
        src_df = read_input_to_df(
            input_file, marketplace_type,
            header_row=general_header_row,
            data_row=general_data_row,
            sheet_name=selected_sheet
        )
    except Exception as e:
        st.error(f"Failed to parse uploaded file: {e}")
        src_df = None

    if src_df is not None:
        if marketplace_type == "General":
            st.markdown("**Sample data (first 3 rows)**")
            st.dataframe(src_df.head(3))
            cols = ["(none)"] + [str(c) for c in src_df.columns]
            col1, col2 = st.columns(2)
            with col1:
                selected_variant_col = st.selectbox("Style Code → productId (leave '(none)' to skip)", options=cols, index=0)
            with col2:
                selected_product_col = st.selectbox("Seller SKU → variantId (leave '(none)' to skip)", options=cols, index=0)
        else:
            if marketplace_type == "Amazon":
                filtered = src_df.attrs.get("filtered_parent_rows", 0)
                if filtered:
                    st.info(f"ℹ️ {filtered} Parent row(s) removed (Parentage Level = 'Parent')")
            st.subheader("Preview (first 5 rows)")
            try:
                st.dataframe(src_df.head(5))
            except Exception as e:
                st.warning(f"Could not render preview: {e}")
                st.write(src_df.head(5).to_string())

    st.markdown("---")

    if marketplace_type == "General":
        if st.button("Generate Output"):
            with st.spinner("Processing…"):
                try:
                    result, assigned_batch_id = process_file(
                        input_file, marketplace_type,
                        selected_variant_col=selected_variant_col,
                        selected_product_col=selected_product_col,
                        general_header_row=general_header_row,
                        general_data_row=general_data_row,
                        general_sheet_name=selected_sheet,
                    )
                    if result:
                        st.success(f"✅ Output Generated! — BatchID assigned: **{assigned_batch_id}**")
                        st.download_button(
                            "📥 Download Output",
                            data=result,
                            file_name=f"output_template_batch{assigned_batch_id}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_button"
                        )
                except Exception as e:
                    st.error(f"Processing failed: {e}")
    else:
        with st.spinner("Processing…"):
            try:
                result, assigned_batch_id = process_file(
                    input_file, marketplace_type,
                    selected_variant_col=None,
                    selected_product_col=None,
                    general_header_row=general_header_row,
                    general_data_row=general_data_row,
                    general_sheet_name=None,
                )
                if result:
                    st.success(f"✅ Output Generated! — BatchID assigned: **{assigned_batch_id}**")
                    st.download_button(
                        "📥 Download Output",
                        data=result,
                        file_name=f"output_template_batch{assigned_batch_id}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_button"
                    )
            except Exception as e:
                st.error(f"Processing failed: {e}")
else:
    st.info("Upload a file to enable header-detection and column selection dropdowns (General only).")

st.markdown("---")
st.caption("Built for Rubick.ai | By Vishnu Sai")
